"""Adaptor for sending Email and SMS
This script allows the user to send Email and SMS with four different services:
1. farazsms for SMS
2. mailchimp for email
3. Twilio for SMS
4. sendgrid for Email
This tool accepts comma separated value files (.csv) as well as excel
(.xls, .xlsx) files.
This script requires that `sendgrid`, `ippanel`, `mailchimp_transactional`
and `twilio`  be installed within the Python environment you are running this script in.
This file can also be imported as a module and contains the following
functions:
    * send - send email and SMS with given content from given sender to given receivers

"""

import codecs
from abc import abstractmethod

import openpyxl as xl
import pandas as pd


class Client:
    """
    Class client(main module) can be used for sending Email and sms by 4 choosing services
    Attributes
    ----------
        file_or_dict: str
            choose to put receiver's information by file_or_dict
            example: {'file' or 'dict'}
        service: str
            choose servise for sending email or sms
            example: {'SendGrid' or 'MailChimp' or 'Twilio' or 'FarazSMS'}
        from_sender: str
            registered sender's email or phone number
        to_receivers: list
            list of receivers => email or phone number
        subject: str
            just emails subject
        content: str
            putting constant content here
        content_path: str
            address of dynamic content
            constant content => __
            content's format: __{token's name}__
            example:
                hello {name},id : {id} phone number {phone_number} email: {email} group {group} promocode: {off}
        data_path: str
            address of tokens file
            supported file format: { excel or csv }
            example:
                    for email ==>   {'mohsen.k9731@gmail.com':
                                        {'name': 'mohsen',
                                        'phone_number': 989101000000,
                                        'email': 'mohsen.k9731@gmail.com',
                                        'off': '10%', 'id': 33446,
                                        'group': 'Candidate'},
                                    }
                    for sms ==>     {'+989399495401':
                                        {'name': 'laya',
                                        'phone_number': 989399495401,
                                        'email': 'laya01@rastava.com',
                                        'off': '30%', 'id': 54123,
                                        'group': 'OrganizationHrAdmin'},
                                    }
        templates: boolean
            for constant content set False
            for dynamic content set True
        tokens: dict
            receivers information
            key of dict = receiver's email or phone number
            value of dict: dict
                key = tokens name
                value = tokens value

    Methods
    -------
        choose_services(argument):
            returns the class of chosen service
        import_from_file():
            read tokens from file
        send():
            call create and send methods of chosen service
        create(): @abstractmethod
            create, set and change attributes that service needs
        create_template(): @abstractmethod
            create Attributes that servise needs for send email with template
        send_with_service(): @abstractmethod
            send email without template (only simple content)
        send_with_template(): @abstractmethod
            send email with template (dynamic content)
        

    """

    def choose_services(self, argument=""):
        """
        returns the class of chosen service
        
        Parameters
        ----------
            argument: str
                name of chosen service

        Returns
        -------
            returns the class of chosen service


        """
        import farazsms_service
        import mailchimp_service
        import sendgrid_service
        import twilio_service
        switcher = {
            'SendGrid': sendgrid_service.sendGrid,
            'MailChimp': mailchimp_service.MailChimp,
            'Twilio': twilio_service.Twilio,
            'FarazSMS': farazsms_service.FarazSMS,
        }
        return switcher.get(argument)

    def __init__(self, file_or_dict="", service="", subject="", from_sender="", to_receivers=[], content="",
                 content_path="", data_path="",
                 templates=False, tokens=""):
        """
        Constructs all the necessary attributes for the client object.

        Parameter
        ---------
            file_or_dict: str
                choose to put receiver's information by file_or_dict
                example: {'file' or 'dict'}
            service: str
                choose servise for sending email or sms
                example: {'SendGrid' or 'MailChimp' or 'Twilio' or 'FarazSMS'}
            from_sender: str
                registered sender's email or phone number
            to_receivers: list
                list of receivers => email or phone number
            subject: str
                just emails subject
            content: str
                putting constant content here
            content_path: str
                address of dynamic content
                constant content => __
                content's format: __{token's name}__
                example:
                    hello {name},id : {id} phone number {phone_number} email: {email} group {group} promocode: {off}
            data_path: str
                address of tokens file
                supported file format: { excel or csv }
                example:
                        for email ==> {'mohsen.k9731@gmail.com':
                                        {'name': 'mohsen',
                                        'phone_number': 989101000000,
                                        'email': 'mohsen.k9731@gmail.com',
                                        'off': '10%', 'id': 33446,
                                        'group': 'Candidate'},
                                      }
                        for sms ==>   {'+989399495401':
                                        {'name': 'laya',
                                        'phone_number': 989399495401,
                                        'email': 'laya01@rastava.com',
                                        'off': '30%', 'id': 54123,
                                        'group': 'OrganizationHrAdmin'},
                                      }
            templates: boolean
                for constant content set False
                for dynamic content set True
            tokens: dict
                receivers information
                key of dict = receiver's email or phone number
                value of dict: dict
                    key = tokens name
                    value = tokens value
            tokens_list: list
                converte list of tokens

        """
        self.subject = subject
        self.content = content
        self.file_or_dict = file_or_dict
        self.service = service
        self.templates = templates
        self.from_sender = from_sender
        self.to_receivers = to_receivers
        self.content_path = content_path
        self.data_path = data_path
        self.tokens = tokens
        self.tokens_list = []

        # check out tokens attribute and assigning to tokens_list
        if (self.tokens != {}) and (self.tokens != ""):
            tokens = []
            for receiver in self.tokens:
                tokens.append(self.tokens[receiver])
            self.tokens_list = tokens

        if self.content_path != "":
            try:
                self.content = codecs.open(self.content_path)
                self.content = self.content.read()
            except:
                raise Exception("Error:FileExistsError")

        if self.file_or_dict == "file":
            Client.import_from_file(self)

    def import_from_file(self):
        """
        convert csv to excel
        assign information to attributes

        """

        if self.data_path.find("csv") != -1:
            df_csv = pd.read_csv(self.data_path)
            self.data_path = self.data_path.replace("csv", "xlsx")
            df_csv.to_excel(self.data_path)

        wb = xl.load_workbook(self.data_path)
        sheet = wb.worksheets[0]

        content = []

        tokens_name = []
        tokens_list = []
        tokens = {}

        firstcell = 1
        if sheet.cell(1, 1).value == None:
            firstcell = 2
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 1).value == None:
                break
            for column in range(firstcell, sheet.max_column + 1):
                if sheet.cell(row, column).value == None:
                    break
                tokens_name.append(sheet.cell(1, column).value)
                value = sheet.cell(row, column).value
                if tokens_name[column - 2] == 'content':
                    content.append(value)
                tokens[tokens_name[column - 2]] = value
            tokens_list.append(tokens)
            tokens = {}

        self.tokens_list = tokens_list

        if content != []:
            self.content = content

    def send(self):
        """
        call create and send methods of chosen service

        """

        if not self.templates:
            self.choose_services(argument=self.service).create(self)
            try:
                self.choose_services(argument=self.service).send_with_service(self)
            except Exception as e:
                print(e)
        else:
            self.choose_services(argument=self.service).create(self)
            self.choose_services(argument=self.service).create_template(self)
            try:
                self.choose_services(argument=self.service).send_with_template(self)
            except Exception as e:
                print(e)

    @abstractmethod
    def create(self):
        raise Exception('Must implement')

    @abstractmethod
    def create_template(self):
        raise Exception('Must implement')

    @abstractmethod
    def send_with_service(self):
        raise Exception('Must implement')

    @abstractmethod
    def send_with_template(self):
        raise Exception('Must implement')
